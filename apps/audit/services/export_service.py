# -*- coding: utf-8 -*-
"""
Export service for generating CSV/Excel exports.
"""
import csv
import io
import logging
from typing import Optional, Dict, Any, List, Type
from datetime import datetime
from django.contrib.auth import get_user_model
from django.db.models import Model, QuerySet
from django.utils import timezone
from django.http import HttpResponse

from ..models import DataExportLog

logger = logging.getLogger(__name__)
User = get_user_model()


class ExportService:
    """Service for data export operations."""
    
    def create_export_log(
        self,
        user: User,
        export_type: str,
        export_format: str,
        filters: Optional[Dict[str, Any]] = None,
    ) -> DataExportLog:
        """Create an export log entry."""
        return DataExportLog.objects.create(
            user=user,
            export_type=export_type,
            export_format=export_format,
            filters=filters or {},
            status=DataExportLog.ExportStatus.PENDING,
        )
    
    def export_to_csv(
        self,
        queryset: QuerySet,
        fields: List[str],
        field_labels: Optional[Dict[str, str]] = None,
        filename: str = 'export.csv',
    ) -> HttpResponse:
        """Export queryset to CSV."""
        field_labels = field_labels or {}
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write header
        header = [field_labels.get(f, f) for f in fields]
        writer.writerow(header)
        
        # Write data
        for obj in queryset:
            row = []
            for field in fields:
                value = self._get_field_value(obj, field)
                row.append(value)
            writer.writerow(row)
        
        return response
    
    def export_to_excel(
        self,
        queryset: QuerySet,
        fields: List[str],
        field_labels: Optional[Dict[str, str]] = None,
        filename: str = 'export.xlsx',
        sheet_name: str = 'Data',
    ) -> HttpResponse:
        """Export queryset to Excel."""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed")
            raise ImportError("openpyxl is required for Excel export")
        
        field_labels = field_labels or {}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write header
        for col, field in enumerate(fields, 1):
            ws.cell(row=1, column=col, value=field_labels.get(field, field))
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = self._get_field_value(obj, field)
                ws.cell(row=row_num, column=col, value=value)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def export_messages(
        self,
        queryset: QuerySet,
        export_format: str = 'csv',
        user: Optional[User] = None,
    ) -> HttpResponse:
        """Export messages."""
        fields = [
            'id', 'whatsapp_message_id', 'direction', 'message_type',
            'status', 'from_number', 'to_number', 'text_body',
            'sent_at', 'delivered_at', 'read_at', 'created_at',
        ]
        field_labels = {
            'id': 'ID',
            'whatsapp_message_id': 'WhatsApp ID',
            'direction': 'Direção',
            'message_type': 'Tipo',
            'status': 'Status',
            'from_number': 'De',
            'to_number': 'Para',
            'text_body': 'Mensagem',
            'sent_at': 'Enviado em',
            'delivered_at': 'Entregue em',
            'read_at': 'Lido em',
            'created_at': 'Criado em',
        }
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if export_format == 'excel':
            return self.export_to_excel(
                queryset, fields, field_labels,
                filename=f'mensagens_{timestamp}.xlsx',
                sheet_name='Mensagens'
            )
        else:
            return self.export_to_csv(
                queryset, fields, field_labels,
                filename=f'mensagens_{timestamp}.csv'
            )
    
    def export_orders(
        self,
        queryset: QuerySet,
        export_format: str = 'csv',
        user: Optional[User] = None,
    ) -> HttpResponse:
        """Export orders."""
        fields = [
            'id', 'order_number', 'store__name', 'store__slug',
            'customer_phone', 'customer_name', 'customer_email',
            'status', 'payment_status', 'payment_method',
            'subtotal', 'discount', 'delivery_fee', 'tax', 'total',
            'delivery_method', 'delivery_address',
            'created_at', 'paid_at', 'shipped_at', 'delivered_at', 'cancelled_at',
        ]
        field_labels = {
            'id': 'ID',
            'order_number': 'Número do Pedido',
            'store__name': 'Loja',
            'store__slug': 'Slug da Loja',
            'customer_phone': 'Telefone',
            'customer_name': 'Nome do Cliente',
            'customer_email': 'Email',
            'status': 'Status',
            'payment_status': 'Status do Pagamento',
            'payment_method': 'Método de Pagamento',
            'subtotal': 'Subtotal',
            'discount': 'Desconto',
            'delivery_fee': 'Taxa de Entrega',
            'tax': 'Impostos',
            'total': 'Total',
            'delivery_method': 'Método de Entrega',
            'delivery_address': 'Endereço de Entrega',
            'created_at': 'Criado em',
            'paid_at': 'Pago em',
            'shipped_at': 'Enviado em',
            'delivered_at': 'Entregue em',
            'cancelled_at': 'Cancelado em',
        }
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if export_format == 'excel':
            return self.export_to_excel(
                queryset, fields, field_labels,
                filename=f'pedidos_{timestamp}.xlsx',
                sheet_name='Pedidos'
            )
        else:
            return self.export_to_csv(
                queryset, fields, field_labels,
                filename=f'pedidos_{timestamp}.csv'
            )
    
    def export_conversations(
        self,
        queryset: QuerySet,
        export_format: str = 'csv',
        user: Optional[User] = None,
    ) -> HttpResponse:
        """Export conversations."""
        fields = [
            'id', 'phone_number', 'contact_name', 'mode', 'status',
            'tags', 'last_message_at', 'created_at', 'closed_at', 'resolved_at',
        ]
        field_labels = {
            'id': 'ID',
            'phone_number': 'Telefone',
            'contact_name': 'Nome do Contato',
            'mode': 'Modo',
            'status': 'Status',
            'tags': 'Tags',
            'last_message_at': 'Última Mensagem',
            'created_at': 'Criado em',
            'closed_at': 'Fechado em',
            'resolved_at': 'Resolvido em',
        }
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if export_format == 'excel':
            return self.export_to_excel(
                queryset, fields, field_labels,
                filename=f'conversas_{timestamp}.xlsx',
                sheet_name='Conversas'
            )
        else:
            return self.export_to_csv(
                queryset, fields, field_labels,
                filename=f'conversas_{timestamp}.csv'
            )
    
    
    def _get_field_value(self, obj: Model, field: str) -> Any:
        """Get field value from object, handling nested fields."""
        parts = field.split('__')
        value = obj
        
        for part in parts:
            if value is None:
                return ''
            value = getattr(value, part, None)
        
        if value is None:
            return ''
        
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        
        if isinstance(value, (list, dict)):
            import json
            return json.dumps(value, ensure_ascii=False)
        
        return str(value)
