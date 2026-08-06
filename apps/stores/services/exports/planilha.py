"""
Planilhas que o dono abre no Excel e entende sem precisar formatar nada.

O export era CSV cru: sem tipo, sem moeda, número com ponto decimal que o Excel
pt-BR lê como texto, e a primeira linha rolando junto com os dados. Aqui a
planilha sai pronta — cabeçalho fixo, filtro, moeda em R$, coluna dimensionada e
linha de totais.

Uso:
    colunas = [
        Coluna('Pedido', 'numero'),
        Coluna('Data', 'data', tipo='data'),
        Coluna('Total', 'total', tipo='dinheiro', somar=True),
    ]
    return resposta_xlsx(linhas, colunas, 'pedidos.xlsx', titulo='Pedidos')
"""
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Callable, Iterable, Optional

from django.http import HttpResponse

# Formatos numéricos do Excel. O ponto-e-vírgula separa positivo;negativo.
FORMATO_DINHEIRO = 'R$ #,##0.00'
FORMATO_INTEIRO = '#,##0'
FORMATO_DECIMAL = '#,##0.00'
FORMATO_PERCENTUAL = '0.0%'
FORMATO_DATA = 'dd/mm/yyyy'
FORMATO_DATA_HORA = 'dd/mm/yyyy hh:mm'

_LARGURA_MIN = 10
_LARGURA_MAX = 60

CONTENT_TYPE_XLSX = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


@dataclass
class Coluna:
    """Uma coluna da planilha.

    `origem` é o nome da chave no dict da linha, ou um callable que recebe a
    linha inteira — útil quando o valor é derivado (ex.: juntar itens do pedido).
    `somar=True` faz a coluna entrar na linha de totais do rodapé.
    """

    titulo: str
    origem: Any  # str (chave) ou Callable[[dict], Any]
    tipo: str = 'texto'  # texto | dinheiro | inteiro | decimal | percentual | data | data_hora
    somar: bool = False
    largura: Optional[int] = None

    def valor(self, linha: dict) -> Any:
        if callable(self.origem):
            return self.origem(linha)
        return linha.get(self.origem)


_FORMATO_POR_TIPO = {
    'dinheiro': FORMATO_DINHEIRO,
    'inteiro': FORMATO_INTEIRO,
    'decimal': FORMATO_DECIMAL,
    'percentual': FORMATO_PERCENTUAL,
    'data': FORMATO_DATA,
    'data_hora': FORMATO_DATA_HORA,
}

_TIPOS_NUMERICOS = {'dinheiro', 'inteiro', 'decimal', 'percentual'}


def _normaliza(valor: Any, tipo: str) -> Any:
    """Converte para um tipo que o Excel entende como número/data, não texto."""
    if valor is None:
        return 0 if tipo in _TIPOS_NUMERICOS else ''
    if tipo in _TIPOS_NUMERICOS:
        if isinstance(valor, Decimal):
            return float(valor)
        if isinstance(valor, (int, float)):
            return valor
        try:
            return float(valor)
        except (TypeError, ValueError):
            return 0
    if tipo in ('data', 'data_hora'):
        if isinstance(valor, datetime):
            # openpyxl não serializa datetime com tzinfo.
            return valor.replace(tzinfo=None) if valor.tzinfo else valor
        if isinstance(valor, date):
            return valor
        return str(valor)
    return '' if valor is None else str(valor)


def _largura_texto(valor, tipo: str) -> int:
    """Largura aproximada da célula COMO ELA APARECE, não do dado cru.

    999.75 tem 6 caracteres mas o Excel exibe "R$ 999,75" (9). Medir o valor
    bruto deixava a coluna de dinheiro estreita e a de percentual larga demais —
    era o motivo de "Receita" sair com 11 e "Preço médio" com 21.
    """
    if valor is None:
        return 0
    if tipo == 'dinheiro':
        return len(f'{float(valor):,.2f}') + 6      # "R$ " + separador de milhar
    if tipo == 'percentual':
        return 9                                    # "100,0%" cabe folgado
    if tipo in ('inteiro', 'decimal'):
        return len(f'{float(valor):,.0f}') + 3
    if tipo == 'data':
        return 12                                   # dd/mm/aaaa
    if tipo == 'data_hora':
        return 18
    return len(str(valor)) + 3


def monta_planilha(
    linhas: Iterable[dict],
    colunas: list[Coluna],
    titulo: str = 'Dados',
    subtitulo: str = '',
):
    """Workbook de uma aba só. Separado da resposta HTTP para poder ser testado."""
    from openpyxl import Workbook

    wb = Workbook()
    _preenche_aba(wb.active, linhas, colunas, titulo=titulo, subtitulo=subtitulo)
    return wb


def _preenche_aba(ws, linhas, colunas, titulo='Dados', subtitulo=''):
    """Escreve uma tabela formatada NA worksheet recebida.

    Preenche direto em vez de montar noutro workbook e copiar célula a célula:
    a cópia perdia mesclagem, bordas, autofiltro e — o pior — o number_format,
    então nas abas secundárias a participação aparecia como
    "0.837016938062098" em vez de 83,7%.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws.title = (titulo[:31] or 'Dados')

    # ── Título e subtítulo ───────────────────────────────────────────────
    ultima_col = get_column_letter(max(len(colunas), 1))
    linha_atual = 1
    if subtitulo:
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = Font(size=15, bold=True, color='1F1A15')
        c.alignment = Alignment(vertical='center')
        ws.merge_cells(f'A1:{ultima_col}1')
        ws.row_dimensions[1].height = 22

        c2 = ws.cell(row=2, column=1, value=subtitulo)
        c2.font = Font(size=9, color='8A7E6C')
        c2.alignment = Alignment(vertical='center', wrap_text=False)
        ws.merge_cells(f'A2:{ultima_col}2')
        ws.row_dimensions[2].height = 14
        linha_atual = 4

    linha_cabecalho = linha_atual
    fundo = PatternFill('solid', start_color='1F1A15')      # carvão da marca
    fonte_cabecalho = Font(bold=True, color='C9A24B', size=10.5)  # dourado
    borda_fina = Side(style='thin', color='E8E4DE')
    zebra = PatternFill('solid', start_color='FAF8F5')

    for i, col in enumerate(colunas, start=1):
        c = ws.cell(row=linha_cabecalho, column=i, value=col.titulo)
        c.fill = fundo
        c.font = fonte_cabecalho
        # Número alinha à direita já no cabeçalho, para a coluna ler como bloco.
        horiz = 'right' if col.tipo in _TIPOS_NUMERICOS else 'left'
        c.alignment = Alignment(horizontal=horiz, vertical='center', wrap_text=True)
    # Altura maior: os títulos de duas palavras vinham espremidos.
    ws.row_dimensions[linha_cabecalho].height = 30

    larguras = [len(c.titulo) + 3 for c in colunas]
    totais = {i: 0.0 for i, c in enumerate(colunas, start=1) if c.somar}

    r = linha_cabecalho
    for indice, linha in enumerate(linhas):
        r += 1
        listrada = indice % 2 == 1
        for i, col in enumerate(colunas, start=1):
            bruto = _normaliza(col.valor(linha), col.tipo)
            celula = ws.cell(row=r, column=i, value=bruto)
            fmt = _FORMATO_POR_TIPO.get(col.tipo)
            if fmt:
                celula.number_format = fmt
            celula.border = Border(bottom=borda_fina)
            celula.alignment = Alignment(
                horizontal='right' if col.tipo in _TIPOS_NUMERICOS else 'left',
                vertical='center',
            )
            # Zebra discreta: ajuda a seguir a linha em tabela larga sem virar
            # enfeite.
            if listrada:
                celula.fill = zebra
            if col.somar and isinstance(bruto, (int, float)):
                totais[i] += bruto
            largura_valor = _largura_texto(bruto, col.tipo)
            if largura_valor > larguras[i - 1]:
                larguras[i - 1] = largura_valor

    ultima_linha_dados = r
    tem_dados = ultima_linha_dados > linha_cabecalho

    # Rodapé de totais — o número que o dono procura primeiro.
    if tem_dados and totais:
        r += 1
        fundo_total = PatternFill('solid', start_color='F0EBE2')
        borda_topo = Side(style='medium', color='C9A24B')
        for i, col in enumerate(colunas, start=1):
            c = ws.cell(row=r, column=i)
            c.fill = fundo_total
            c.border = Border(top=borda_topo)
            c.font = Font(bold=True, size=10.5, color='1F1A15')
            if i == 1:
                c.value = 'TOTAL'
            elif i in totais:
                c.value = totais[i]
                fmt = _FORMATO_POR_TIPO.get(col.tipo)
                if fmt:
                    c.number_format = fmt
            c.alignment = Alignment(
                horizontal='right' if col.tipo in _TIPOS_NUMERICOS and i != 1 else 'left',
                vertical='center',
            )
        ws.row_dimensions[r].height = 20

    for i, col in enumerate(colunas, start=1):
        largura = col.largura or min(max(larguras[i - 1], _LARGURA_MIN), _LARGURA_MAX)
        ws.column_dimensions[get_column_letter(i)].width = largura

    # Cabeçalho fixo ao rolar + filtro por coluna. É o que transforma um dump
    # numa planilha utilizável de verdade.
    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)
    if tem_dados:
        ws.auto_filter.ref = (
            f'A{linha_cabecalho}:'
            f'{ultima_col}{ultima_linha_dados}'
        )
    ws.sheet_view.showGridLines = False
    return ws


def resposta_xlsx(
    linhas: Iterable[dict],
    colunas: list[Coluna],
    nome_arquivo: str,
    titulo: str = 'Dados',
    subtitulo: str = '',
    abas_extras: Optional[list[tuple[str, Iterable[dict], list[Coluna]]]] = None,
) -> HttpResponse:
    """Workbook pronto como download. `abas_extras` acrescenta outras planilhas."""
    import io

    wb = monta_planilha(linhas, colunas, titulo=titulo, subtitulo=subtitulo)

    if abas_extras:
        for nome_aba, linhas_extra, colunas_extra in abas_extras:
            _preenche_aba(
                wb.create_sheet(title=nome_aba[:31]),
                linhas_extra, colunas_extra,
                titulo=nome_aba, subtitulo=subtitulo,
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resposta = HttpResponse(buffer.read(), content_type=CONTENT_TYPE_XLSX)
    resposta['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return resposta
