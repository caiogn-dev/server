"""
Importa a tabela TACO (NEPA/Unicamp) inteira para NutritionIngredient.

Antes disso o cadastro tinha 37 alimentos TACO digitados à mão — o lojista
gastava mais tempo procurando o valor do arroz do que montando a receita.

Duas planilhas do mesmo arquivo:
  - "CMVCol taco3": composição centesimal (energia, macros, fibra, sódio)
  - "AGtaco3": ácidos graxos, de onde sai a gordura saturada — que a TACO
    não põe na primeira aba e é justamente um dos 3 gatilhos da lupa frontal
    da RDC 429/2020.

O que a TACO NÃO tem, e por isso fica nulo (não zero):
  - açúcares totais e adicionados: só existem em rótulo de fabricante
  - gordura trans: a 4ª edição não traz a coluna
Nulo é honesto — zero aqui viraria etiqueta afirmando "0 g de açúcar".

Alergênicos NÃO são preenchidos: a TACO não traz esse dado e adivinhar pelo
nome do alimento é justamente o erro que manda alguém para o hospital. Os
importados entram com `allergens_reviewed=False`, e a declaração da receita
se recusa a sair enquanto houver ingrediente não revisado.

Usage:
    python manage.py import_taco_table --arquivo /tmp/taco.xlsx
    python manage.py import_taco_table --arquivo /tmp/taco.xlsx --dry-run
"""
import logging
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.nutrition.models import NutritionIngredient

logger = logging.getLogger(__name__)

ABA_COMPOSICAO = "CMVCol taco3"
ABA_ACIDOS_GRAXOS = "AGtaco3"

# Índice de coluna (0-based) na aba de composição.
COLUNAS = {
    "numero": 0,
    "descricao": 1,
    "energy_kcal": 3,
    "protein_g": 5,
    "total_fat_g": 6,
    "carbohydrates_g": 8,
    "fiber_g": 9,
    "sodium_mg": 17,
}
COL_SATURADA = 2  # na aba de ácidos graxos

# A TACO usa marcadores textuais em vez de vazio.
#   "NA"  = não analisado    -> nulo, ninguém mediu
#   "Tr"  = traço            -> zero, medido e desprezível
NAO_ANALISADO = {"na", "nd", "*", ""}
TRACO = {"tr"}


def _numero(bruto):
    """Converte célula da TACO em Decimal, respeitando NA e Tr."""
    if bruto is None:
        return None
    texto = str(bruto).strip()
    if texto.lower() in NAO_ANALISADO:
        return None
    if texto.lower() in TRACO:
        return Decimal("0")
    try:
        return Decimal(texto.replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _canonico(descricao):
    """Chave estável de busca: minúscula, sem acento sobrando, sem espaço duplo."""
    texto = re.sub(r"\s+", " ", str(descricao)).strip().lower()
    return texto[:255]


def _categoria_da_linha(valor):
    """Linhas de cabeçalho da TACO trazem a categoria na 1ª coluna, sem número."""
    texto = str(valor).strip()
    return texto if texto and not texto.isdigit() else ""


class Command(BaseCommand):
    help = "Importa a tabela TACO completa para o cadastro de ingredientes"

    def add_arguments(self, parser):
        parser.add_argument("--arquivo", required=True, help="Caminho do .xlsx da TACO")
        parser.add_argument("--edicao", default="TACO 4ª edição (NEPA/Unicamp)",
                            help="Texto gravado em source_edition")
        parser.add_argument("--dry-run", action="store_true",
                            help="Conta o que entraria, sem gravar")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl não instalado no ambiente.")

        try:
            wb = openpyxl.load_workbook(options["arquivo"], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Arquivo não encontrado: {options['arquivo']}")

        faltando = [aba for aba in (ABA_COMPOSICAO, ABA_ACIDOS_GRAXOS) if aba not in wb.sheetnames]
        if faltando:
            raise CommandError(f"Planilha sem as abas esperadas: {', '.join(faltando)}. "
                               f"Abas encontradas: {', '.join(wb.sheetnames)}")

        saturada_por_numero = self._ler_saturadas(wb[ABA_ACIDOS_GRAXOS])
        self.stdout.write(f"🧈 {len(saturada_por_numero)} alimentos com gordura saturada na aba de ácidos graxos")

        registros = self._ler_composicao(wb[ABA_COMPOSICAO], saturada_por_numero)
        self.stdout.write(f"📋 {len(registros)} alimentos lidos da composição centesimal")

        if options["dry_run"]:
            for r in registros[:5]:
                self.stdout.write(f"   ex.: [{r['source_code']}] {r['display_name'][:60]} — "
                                  f"{r['energy_kcal']} kcal, sódio {r['sodium_mg']} mg")
            self.stdout.write(self.style.WARNING("--dry-run: nada gravado."))
            return

        criados, atualizados = self._gravar(registros, options["edicao"])

        self.stdout.write(self.style.SUCCESS(f"\n✅ {criados} criados, {atualizados} atualizados"))
        self.stdout.write(self.style.WARNING(
            "⚠️  Alergênicos NÃO foram preenchidos (a TACO não tem esse dado). "
            "Os importados ficam com allergens_reviewed=False e a etiqueta se "
            "recusa a declarar alergênico até alguém revisar."))

    def _ler_saturadas(self, ws):
        """numero do alimento -> gordura saturada em g/100g."""
        saturadas = {}
        for linha in ws.iter_rows(min_row=1, values_only=True):
            if not linha or linha[0] is None:
                continue
            numero = str(linha[0]).strip()
            if not numero.isdigit():
                continue
            valor = _numero(linha[COL_SATURADA]) if len(linha) > COL_SATURADA else None
            if valor is not None:
                saturadas[numero] = valor
        return saturadas

    def _ler_composicao(self, ws, saturada_por_numero):
        registros = []
        categoria_atual = ""
        for linha in ws.iter_rows(min_row=1, values_only=True):
            if not linha or linha[COLUNAS["numero"]] is None:
                continue
            numero = str(linha[COLUNAS["numero"]]).strip()
            if not numero.isdigit():
                # Cabeçalho de grupo ("Cereais e derivados") — vira a categoria
                # dos alimentos que vêm depois.
                categoria = _categoria_da_linha(numero)
                if categoria and len(categoria) < 60:
                    categoria_atual = categoria
                continue

            descricao = linha[COLUNAS["descricao"]]
            if not descricao:
                continue

            registros.append({
                "source_code": numero,
                "canonical_name": _canonico(descricao),
                "display_name": str(descricao).strip()[:255],
                "category": categoria_atual[:80],
                "energy_kcal": _numero(linha[COLUNAS["energy_kcal"]]),
                "protein_g": _numero(linha[COLUNAS["protein_g"]]),
                "total_fat_g": _numero(linha[COLUNAS["total_fat_g"]]),
                "carbohydrates_g": _numero(linha[COLUNAS["carbohydrates_g"]]),
                "fiber_g": _numero(linha[COLUNAS["fiber_g"]]),
                "sodium_mg": _numero(linha[COLUNAS["sodium_mg"]]),
                "saturated_fat_g": saturada_por_numero.get(numero),
                # A TACO não mede: fica nulo, e o cálculo avisa que faltou.
                "total_sugars_g": None,
                "added_sugars_g": None,
                "trans_fat_g": None,
            })
        return registros

    @transaction.atomic
    def _gravar(self, registros, edicao):
        criados = atualizados = 0
        for r in registros:
            # Os 33 TACO digitados à mão antes deste comando têm o nome em
            # Title Case ("Arroz, tipo 1, cozido"). Casar sem diferenciar caixa
            # atualiza aquela linha em vez de criar uma irmã quase idêntica —
            # duas "arroz cozido" na busca é pior que nenhuma.
            existente = NutritionIngredient.objects.filter(
                store__isnull=True, preparation_state="",
                canonical_name__iexact=r["canonical_name"],
            ).first()

            # store=None: a TACO é base pública, compartilhada por todas as lojas.
            campos = {
                    "store": None,
                    "canonical_name": r["canonical_name"],
                    "preparation_state": "",
                    "display_name": r["display_name"],
                    "category": r["category"],
                    "default_unit": "g",
                    "source": NutritionIngredient.Source.TACO,
                    "source_code": r["source_code"],
                    "source_edition": edicao,
                    "energy_kcal": r["energy_kcal"],
                    "protein_g": r["protein_g"],
                    "total_fat_g": r["total_fat_g"],
                    "saturated_fat_g": r["saturated_fat_g"],
                    "trans_fat_g": r["trans_fat_g"],
                    "carbohydrates_g": r["carbohydrates_g"],
                    "total_sugars_g": r["total_sugars_g"],
                    "added_sugars_g": r["added_sugars_g"],
                    "fiber_g": r["fiber_g"],
                    "sodium_mg": r["sodium_mg"],
                    "is_active": True,
            }

            if existente:
                for campo, valor in campos.items():
                    setattr(existente, campo, valor)
                existente.save(update_fields=list(campos))
                atualizados += 1
            else:
                NutritionIngredient.objects.create(**campos)
                criados += 1
        return criados, atualizados
