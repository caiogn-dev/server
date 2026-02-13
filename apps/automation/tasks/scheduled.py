"""
Celery tasks for scheduled messages and automated reports.
"""
import logging
import os
import time
from datetime import datetime, timedelta
from typing import Dict, Any

from celery import shared_task
from django.utils import timezone
from django.conf import settings
from django.core.mail import EmailMessage
from django.db import models

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def send_scheduled_message(self, message_id: str):
    """Send a scheduled message."""
    from ..models import ScheduledMessage
    from apps.whatsapp.services import MessageService
    
    try:
        message = ScheduledMessage.objects.select_related('account').get(
            id=message_id,
            is_active=True
        )
        
        # Check if already processed
        if message.status != ScheduledMessage.Status.PENDING:
            logger.info(f"Scheduled message {message_id} already processed: {message.status}")
            return
        
        # Mark as processing
        message.status = ScheduledMessage.Status.PROCESSING
        message.save(update_fields=['status'])
        
        # Send message
        service = MessageService()
        result = None
        
        if message.message_type == ScheduledMessage.MessageType.TEXT:
            result = service.send_text_message(
                account_id=str(message.account_id),
                to=message.to_number,
                text=message.message_text
            )
        
        elif message.message_type == ScheduledMessage.MessageType.TEMPLATE:
            result = service.send_template_message(
                account_id=str(message.account_id),
                to=message.to_number,
                template_name=message.template_name,
                language_code=message.template_language,
                components=message.template_components
            )
        
        elif message.message_type == ScheduledMessage.MessageType.IMAGE:
            result = service.send_image(
                account_id=str(message.account_id),
                to=message.to_number,
                image_url=message.media_url,
                caption=message.message_text
            )
        
        elif message.message_type == ScheduledMessage.MessageType.DOCUMENT:
            result = service.send_document(
                account_id=str(message.account_id),
                to=message.to_number,
                document_url=message.media_url,
                caption=message.message_text
            )
        
        elif message.message_type == ScheduledMessage.MessageType.INTERACTIVE:
            result = service.send_interactive_buttons(
                account_id=str(message.account_id),
                to=message.to_number,
                body_text=message.message_text,
                buttons=message.buttons
            )
        
        # Update status
        if result:
            message.status = ScheduledMessage.Status.SENT
            message.sent_at = timezone.now()
            message.whatsapp_message_id = result.get('whatsapp_message_id', '')
            logger.info(f"Scheduled message {message_id} sent successfully")
            
            # Send WebSocket notification
            from ..consumers import notify_scheduled_message_sent
            notify_scheduled_message_sent({
                'id': str(message.id),
                'to_number': message.to_number,
                'status': 'sent',
                'sent_at': message.sent_at.isoformat()
            }, str(message.account_id))
        else:
            message.status = ScheduledMessage.Status.FAILED
            message.error_message = "Failed to send message"
            logger.error(f"Scheduled message {message_id} failed to send")
        
        message.save()
        
    except ScheduledMessage.DoesNotExist:
        logger.warning(f"Scheduled message not found: {message_id}")
    except Exception as e:
        logger.error(f"Error sending scheduled message {message_id}: {str(e)}")
        try:
            message = ScheduledMessage.objects.get(id=message_id)
            message.status = ScheduledMessage.Status.FAILED
            message.error_message = str(e)
            message.save()
        except ScheduledMessage.DoesNotExist:
            logger.warning(f"Could not update failed message status - message {message_id} not found")
        raise self.retry(exc=e, countdown=60)


@shared_task
def process_scheduled_messages():
    """
    Process pending scheduled messages.
    Run every minute.
    """
    from ..models import ScheduledMessage
    
    now = timezone.now()
    
    # Find messages that should be sent
    pending_messages = ScheduledMessage.objects.filter(
        status=ScheduledMessage.Status.PENDING,
        scheduled_at__lte=now,
        is_active=True
    )[:100]  # Process max 100 at a time
    
    for message in pending_messages:
        send_scheduled_message.delay(str(message.id))
    
    if pending_messages:
        logger.info(f"Queued {len(pending_messages)} scheduled messages for sending")


@shared_task(bind=True, max_retries=2)
def generate_report(self, schedule_id: str = None, report_type: str = None, 
                    period_start: str = None, period_end: str = None,
                    account_id: str = None, company_id: str = None,
                    recipients: list = None, export_format: str = 'xlsx',
                    user_id: int = None):
    """Generate a report."""
    from ..models import ReportSchedule, GeneratedReport
    from apps.whatsapp.models import Message, WhatsAppAccount
    from apps.conversations.models import Conversation
    from apps.stores.models import StoreOrder, StoreIntegration
    from apps.automation.models import CustomerSession, AutomationLog
    
    start_time = time.time()
    
    try:
        # Get schedule if provided
        schedule = None
        if schedule_id:
            schedule = ReportSchedule.objects.get(id=schedule_id, is_active=True)
            report_type = schedule.report_type
            account_id = str(schedule.account_id) if schedule.account_id else None
            company_id = str(schedule.company_id) if schedule.company_id else None
            recipients = schedule.recipients
            export_format = schedule.export_format
        
        # Parse dates
        if period_start:
            start = datetime.fromisoformat(period_start.replace('Z', '+00:00'))
        else:
            start = timezone.now() - timedelta(days=7)
        
        if period_end:
            end = datetime.fromisoformat(period_end.replace('Z', '+00:00'))
        else:
            end = timezone.now()
        
        # Create report record
        report = GeneratedReport.objects.create(
            schedule=schedule,
            name=f"Report_{report_type}_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}",
            report_type=report_type,
            period_start=start,
            period_end=end,
            file_format=export_format,
            created_by_id=user_id
        )
        
        # Generate report data
        data = {}
        records_count = 0
        
        store_ids = None
        if account_id:
            try:
                account = WhatsAppAccount.objects.get(id=account_id)
                integrations = StoreIntegration.objects.filter(
                    integration_type=StoreIntegration.IntegrationType.WHATSAPP,
                ).filter(
                    models.Q(phone_number_id=account.phone_number_id) |
                    models.Q(waba_id=account.waba_id)
                ).values_list('store_id', flat=True)
                store_ids = list(integrations)
            except Exception:
                store_ids = []

        if report_type in ['messages', 'full']:
            queryset = Message.objects.filter(created_at__gte=start, created_at__lte=end)
            if account_id:
                queryset = queryset.filter(account_id=account_id)
            data['messages'] = list(queryset.values(
                'id', 'direction', 'message_type', 'status', 'from_number', 
                'to_number', 'text_body', 'created_at'
            )[:5000])
            records_count += len(data['messages'])
        
        if report_type in ['orders', 'full']:
            queryset = StoreOrder.objects.select_related('store').filter(
                created_at__gte=start,
                created_at__lte=end
            )
            if store_ids is not None:
                queryset = queryset.filter(store_id__in=store_ids)
            data['orders'] = list(queryset.values(
                'id', 'order_number', 'customer_phone', 'customer_name',
                'status', 'payment_status', 'payment_method', 'total',
                'store__name', 'store__slug', 'created_at'
            )[:5000])
            records_count += len(data['orders'])
        
        if report_type in ['conversations', 'full']:
            queryset = Conversation.objects.filter(created_at__gte=start, created_at__lte=end)
            if account_id:
                queryset = queryset.filter(account_id=account_id)
            data['conversations'] = list(queryset.values(
                'id', 'phone_number', 'contact_name', 'mode', 'status', 
                'last_message_at', 'created_at'
            )[:5000])
            records_count += len(data['conversations'])
        
        if report_type in ['payments', 'full']:
            queryset = StoreOrder.objects.select_related('store').filter(
                created_at__gte=start,
                created_at__lte=end
            )
            if store_ids is not None:
                queryset = queryset.filter(store_id__in=store_ids)
            data['payments'] = list(queryset.values(
                'id', 'order_number', 'payment_status', 'payment_method',
                'total', 'paid_at', 'created_at', 'store__name', 'store__slug'
            )[:5000])
            records_count += len(data['payments'])
        
        if report_type in ['automation', 'full']:
            if company_id:
                sessions = CustomerSession.objects.filter(
                    company_id=company_id,
                    created_at__gte=start, 
                    created_at__lte=end
                )
                logs = AutomationLog.objects.filter(
                    company_id=company_id,
                    created_at__gte=start, 
                    created_at__lte=end
                )
            else:
                sessions = CustomerSession.objects.filter(
                    created_at__gte=start, 
                    created_at__lte=end
                )
                logs = AutomationLog.objects.filter(
                    created_at__gte=start, 
                    created_at__lte=end
                )
            
            data['sessions'] = list(sessions.values(
                'id', 'phone_number', 'customer_name', 'status', 
                'cart_total', 'created_at'
            )[:5000])
            data['automation_logs'] = list(logs.values(
                'id', 'action_type', 'description', 'phone_number', 
                'is_error', 'created_at'
            )[:5000])
            records_count += len(data.get('sessions', [])) + len(data.get('automation_logs', []))
        
        # Generate file
        file_path = _generate_report_file(report, data, export_format)
        
        # Update report
        generation_time = int((time.time() - start_time) * 1000)
        report.status = GeneratedReport.Status.COMPLETED
        report.file_path = file_path
        report.file_size = os.path.getsize(file_path) if file_path and os.path.exists(file_path) else 0
        report.records_count = records_count
        report.generation_time_ms = generation_time
        report.save()
        
        # Send email if recipients provided
        if recipients:
            _send_report_email(report, recipients)
        
        # Update schedule if applicable
        if schedule:
            schedule.last_run_at = timezone.now()
            schedule.run_count += 1
            schedule.last_error = ''
            schedule.calculate_next_run()
            schedule.save()
        
        # Send WebSocket notification
        from ..consumers import notify_report_generated
        notify_report_generated({
            'id': str(report.id),
            'name': report.name,
            'status': 'completed',
            'records_count': records_count
        })
        
        logger.info(f"Report {report.id} generated successfully with {records_count} records")
        return str(report.id)
        
    except ReportSchedule.DoesNotExist:
        logger.warning(f"Report schedule not found: {schedule_id}")
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        if 'report' in locals():
            report.status = GeneratedReport.Status.FAILED
            report.error_message = str(e)
            report.save()
        if schedule_id:
            try:
                schedule = ReportSchedule.objects.get(id=schedule_id)
                schedule.last_error = str(e)
                schedule.save()
            except ReportSchedule.DoesNotExist:
                logger.warning(f"Could not update schedule error - schedule {schedule_id} not found")
        raise self.retry(exc=e, countdown=300)


def _generate_report_file(report, data: Dict[str, Any], export_format: str) -> str:
    """Generate report file and return path."""
    import json
    
    # Create reports directory
    reports_dir = os.path.join(settings.BASE_DIR, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"report_{report.report_type}_{timestamp}"
    
    if export_format == 'xlsx':
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            for sheet_name, sheet_data in data.items():
                if not sheet_data:
                    continue
                
                ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit
                
                # Write headers
                if sheet_data:
                    headers = list(sheet_data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Write data
                    for row_idx, row_data in enumerate(sheet_data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            value = row_data.get(header, '')
                            if isinstance(value, datetime):
                                value = value.isoformat()
                            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            
            file_path = os.path.join(reports_dir, f"{filename}.xlsx")
            wb.save(file_path)
            return file_path
            
        except ImportError:
            # Fallback to CSV
            export_format = 'csv'
    
    if export_format == 'csv':
        import csv
        
        file_path = os.path.join(reports_dir, f"{filename}.csv")
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = None
            for sheet_name, sheet_data in data.items():
                if not sheet_data:
                    continue
                
                # Write section header
                f.write(f"\n=== {sheet_name.upper()} ===\n")
                
                headers = list(sheet_data[0].keys())
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                for row in sheet_data:
                    writer.writerow({k: str(v) if v else '' for k, v in row.items()})
        
        return file_path
    
    return ''


def _send_report_email(report, recipients: list):
    """Send report via email."""
    from django.core.mail import EmailMessage
    
    try:
        subject = f"Relatório: {report.name}"
        body = f"""
Olá,

Seu relatório foi gerado com sucesso.

Detalhes:
- Tipo: {report.report_type}
- Período: {report.period_start.strftime('%d/%m/%Y')} a {report.period_end.strftime('%d/%m/%Y')}
- Registros: {report.records_count}
- Tempo de geração: {report.generation_time_ms}ms

O arquivo está anexado a este email.

Atenciosamente,
Sistema WhatsApp Business
        """
        
        email = EmailMessage(
            subject=subject,
            body=body,
            to=recipients
        )
        
        if report.file_path and os.path.exists(report.file_path):
            email.attach_file(report.file_path)
        
        email.send()
        
        report.email_sent = True
        report.email_sent_at = timezone.now()
        report.email_recipients = recipients
        report.save()
        
        logger.info(f"Report email sent to {recipients}")
        
    except Exception as e:
        logger.error(f"Failed to send report email: {str(e)}")


@shared_task
def process_scheduled_reports():
    """
    Process scheduled reports that are due.
    Run every hour.
    """
    from ..models import ReportSchedule
    
    now = timezone.now()
    
    # Find reports that should run
    due_reports = ReportSchedule.objects.filter(
        status=ReportSchedule.Status.ACTIVE,
        next_run_at__lte=now,
        is_active=True
    )
    
    for schedule in due_reports:
        generate_report.delay(schedule_id=str(schedule.id))
        logger.info(f"Queued report generation for schedule: {schedule.name}")


@shared_task
def cleanup_old_reports():
    """
    Clean up old generated reports.
    Run daily.
    """
    from ..models import GeneratedReport
    
    # Delete reports older than 30 days
    threshold = timezone.now() - timedelta(days=30)
    
    old_reports = GeneratedReport.objects.filter(created_at__lt=threshold)
    
    for report in old_reports:
        # Delete file
        if report.file_path and os.path.exists(report.file_path):
            try:
                os.remove(report.file_path)
            except OSError as e:
                logger.warning(f"Failed to delete report file {report.file_path}: {str(e)}")
        
        report.delete()
    
    logger.info(f"Cleaned up {old_reports.count()} old reports")
