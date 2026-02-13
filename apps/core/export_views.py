"""
Data Export Views - Export data to CSV/Excel formats.
"""
import csv
import io
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any

from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiParameter

logger = logging.getLogger(__name__)


class Echo:
    """An object that implements just the write method of the file-like interface."""
    def write(self, value):
        return value


def generate_csv_response(data: List[Dict], filename: str, fieldnames: List[str] = None):
    """Generate a streaming CSV response."""
    if not data:
        return HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    
    if not fieldnames:
        fieldnames = list(data[0].keys())
    
    pseudo_buffer = Echo()
    writer = csv.DictWriter(pseudo_buffer, fieldnames=fieldnames, extrasaction='ignore')
    
    def generate():
        yield writer.writeheader()
        for row in data:
            yield writer.writerow(row)
    
    response = StreamingHttpResponse(
        generate(),
        content_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
    return response


def generate_excel_response(data: List[Dict], filename: str, sheet_name: str = 'Data'):
    """Generate an Excel response using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl is not installed
        return generate_csv_response(data, filename.replace('.xlsx', '.csv'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not data:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    
    # Write headers
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            # Convert datetime objects to strings
            if isinstance(value, datetime):
                value = value.isoformat()
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@extend_schema(
    parameters=[
        OpenApiParameter(name='account_id', type=str, required=False),
        OpenApiParameter(name='start_date', type=str, required=False),
        OpenApiParameter(name='end_date', type=str, required=False),
        OpenApiParameter(name='format', type=str, required=False, enum=['csv', 'xlsx']),
        OpenApiParameter(name='direction', type=str, required=False, enum=['inbound', 'outbound']),
        OpenApiParameter(name='status', type=str, required=False),
    ],
    tags=['Export']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_messages(request):
    """Export messages to CSV or Excel."""
    from apps.whatsapp.models import Message
    
    # Get parameters
    account_id = request.query_params.get('account_id')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'csv')
    direction = request.query_params.get('direction')
    message_status = request.query_params.get('status')
    
    # Build queryset
    queryset = Message.objects.select_related('account').all()
    
    if account_id:
        queryset = queryset.filter(account_id=account_id)
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__lte=end)
        except ValueError:
            pass
    
    if direction:
        queryset = queryset.filter(direction=direction)
    
    if message_status:
        queryset = queryset.filter(status=message_status)
    
    # Limit to 10000 records
    queryset = queryset.order_by('-created_at')[:10000]
    
    # Prepare data
    data = []
    for msg in queryset:
        data.append({
            'id': str(msg.id),
            'account_name': msg.account.name if msg.account else '',
            'whatsapp_message_id': msg.whatsapp_message_id,
            'direction': msg.direction,
            'message_type': msg.message_type,
            'status': msg.status,
            'from_number': msg.from_number,
            'to_number': msg.to_number,
            'text_body': msg.text_body or '',
            'template_name': msg.template_name or '',
            'error_code': msg.error_code or '',
            'error_message': msg.error_message or '',
            'sent_at': msg.sent_at.isoformat() if msg.sent_at else '',
            'delivered_at': msg.delivered_at.isoformat() if msg.delivered_at else '',
            'read_at': msg.read_at.isoformat() if msg.read_at else '',
            'created_at': msg.created_at.isoformat(),
        })
    
    # Generate response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if export_format == 'xlsx':
        return generate_excel_response(data, f'messages_{timestamp}.xlsx', 'Messages')
    return generate_csv_response(data, f'messages_{timestamp}.csv')


@extend_schema(
    parameters=[
        OpenApiParameter(name='account_id', type=str, required=False),
        OpenApiParameter(name='store', type=str, required=False),
        OpenApiParameter(name='start_date', type=str, required=False),
        OpenApiParameter(name='end_date', type=str, required=False),
        OpenApiParameter(name='format', type=str, required=False, enum=['csv', 'xlsx']),
        OpenApiParameter(name='status', type=str, required=False),
    ],
    tags=['Export']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_orders(request):
    """Export orders to CSV or Excel."""
    from apps.stores.models import StoreOrder
    
    # Get parameters
    account_id = request.query_params.get('account_id')
    store_param = request.query_params.get('store')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'csv')
    order_status = request.query_params.get('status')
    
    # Build queryset
    queryset = StoreOrder.objects.select_related('store').all()
    if store_param:
        try:
            import uuid as uuid_module
            uuid_module.UUID(store_param)
            queryset = queryset.filter(store_id=store_param)
        except (ValueError, AttributeError):
            queryset = queryset.filter(store__slug=store_param)
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__lte=end)
        except ValueError:
            pass
    
    if order_status:
        queryset = queryset.filter(status=order_status)
    
    # Limit to 10000 records
    queryset = queryset.order_by('-created_at')[:10000]
    
    # Prepare data
    data = []
    for order in queryset:
        data.append({
            'id': str(order.id),
            'order_number': order.order_number,
            'store_name': order.store.name if order.store else '',
            'store_slug': order.store.slug if order.store else '',
            'customer_phone': order.customer_phone,
            'customer_name': order.customer_name,
            'customer_email': order.customer_email,
            'status': order.status,
            'payment_status': order.payment_status,
            'payment_method': order.payment_method,
            'subtotal': float(order.subtotal),
            'discount': float(order.discount),
            'delivery_fee': float(order.delivery_fee),
            'tax': float(order.tax),
            'total': float(order.total),
            'currency': order.store.currency if order.store else '',
            'customer_notes': order.customer_notes or '',
            'internal_notes': order.internal_notes or '',
            'paid_at': order.paid_at.isoformat() if order.paid_at else '',
            'shipped_at': order.shipped_at.isoformat() if order.shipped_at else '',
            'delivered_at': order.delivered_at.isoformat() if order.delivered_at else '',
            'cancelled_at': order.cancelled_at.isoformat() if order.cancelled_at else '',
            'created_at': order.created_at.isoformat(),
        })
    
    # Generate response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if export_format == 'xlsx':
        return generate_excel_response(data, f'orders_{timestamp}.xlsx', 'Orders')
    return generate_csv_response(data, f'orders_{timestamp}.csv')


@extend_schema(
    parameters=[
        OpenApiParameter(name='company_id', type=str, required=False),
        OpenApiParameter(name='start_date', type=str, required=False),
        OpenApiParameter(name='end_date', type=str, required=False),
        OpenApiParameter(name='format', type=str, required=False, enum=['csv', 'xlsx']),
        OpenApiParameter(name='status', type=str, required=False),
    ],
    tags=['Export']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_sessions(request):
    """Export customer sessions to CSV or Excel."""
    from apps.automation.models import CustomerSession
    
    # Get parameters
    company_id = request.query_params.get('company_id')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'csv')
    session_status = request.query_params.get('status')
    
    # Build queryset
    queryset = CustomerSession.objects.select_related('company').all()
    
    if company_id:
        queryset = queryset.filter(company_id=company_id)
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__lte=end)
        except ValueError:
            pass
    
    if session_status:
        queryset = queryset.filter(status=session_status)
    
    # Limit to 10000 records
    queryset = queryset.order_by('-created_at')[:10000]
    
    # Prepare data
    data = []
    for session in queryset:
        data.append({
            'id': str(session.id),
            'company_name': session.company.company_name if session.company else '',
            'phone_number': session.phone_number,
            'customer_name': session.customer_name,
            'customer_email': session.customer_email,
            'session_id': session.session_id,
            'status': session.status,
            'cart_total': float(session.cart_total),
            'cart_items_count': session.cart_items_count,
            'external_order_id': session.external_order_id or '',
            'last_activity_at': session.last_activity_at.isoformat() if session.last_activity_at else '',
            'created_at': session.created_at.isoformat(),
        })
    
    # Generate response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if export_format == 'xlsx':
        return generate_excel_response(data, f'sessions_{timestamp}.xlsx', 'Sessions')
    return generate_csv_response(data, f'sessions_{timestamp}.csv')


@extend_schema(
    parameters=[
        OpenApiParameter(name='company_id', type=str, required=False),
        OpenApiParameter(name='start_date', type=str, required=False),
        OpenApiParameter(name='end_date', type=str, required=False),
        OpenApiParameter(name='format', type=str, required=False, enum=['csv', 'xlsx']),
        OpenApiParameter(name='action_type', type=str, required=False),
        OpenApiParameter(name='is_error', type=bool, required=False),
    ],
    tags=['Export']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_automation_logs(request):
    """Export automation logs to CSV or Excel."""
    from apps.automation.models import AutomationLog
    
    # Get parameters
    company_id = request.query_params.get('company_id')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'csv')
    action_type = request.query_params.get('action_type')
    is_error = request.query_params.get('is_error')
    
    # Build queryset
    queryset = AutomationLog.objects.select_related('company').all()
    
    if company_id:
        queryset = queryset.filter(company_id=company_id)
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__lte=end)
        except ValueError:
            pass
    
    if action_type:
        queryset = queryset.filter(action_type=action_type)
    
    if is_error is not None:
        queryset = queryset.filter(is_error=is_error.lower() == 'true')
    
    # Limit to 10000 records
    queryset = queryset.order_by('-created_at')[:10000]
    
    # Prepare data
    data = []
    for log in queryset:
        data.append({
            'id': str(log.id),
            'company_name': log.company.company_name if log.company else '',
            'action_type': log.action_type,
            'description': log.description,
            'phone_number': log.phone_number,
            'event_type': log.event_type,
            'is_error': log.is_error,
            'error_message': log.error_message or '',
            'created_at': log.created_at.isoformat(),
        })
    
    # Generate response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if export_format == 'xlsx':
        return generate_excel_response(data, f'automation_logs_{timestamp}.xlsx', 'Logs')
    return generate_csv_response(data, f'automation_logs_{timestamp}.csv')


@extend_schema(
    parameters=[
        OpenApiParameter(name='account_id', type=str, required=False),
        OpenApiParameter(name='start_date', type=str, required=False),
        OpenApiParameter(name='end_date', type=str, required=False),
        OpenApiParameter(name='format', type=str, required=False, enum=['csv', 'xlsx']),
        OpenApiParameter(name='status', type=str, required=False),
        OpenApiParameter(name='mode', type=str, required=False),
    ],
    tags=['Export']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_conversations(request):
    """Export conversations to CSV or Excel."""
    from apps.conversations.models import Conversation
    
    # Get parameters
    account_id = request.query_params.get('account_id')
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    export_format = request.query_params.get('format', 'csv')
    conv_status = request.query_params.get('status')
    mode = request.query_params.get('mode')
    
    # Build queryset
    queryset = Conversation.objects.select_related('account').all()
    
    if account_id:
        queryset = queryset.filter(account_id=account_id)
    
    if start_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__gte=start)
        except ValueError:
            pass
    
    if end_date:
        try:
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            queryset = queryset.filter(created_at__lte=end)
        except ValueError:
            pass
    
    if conv_status:
        queryset = queryset.filter(status=conv_status)
    
    if mode:
        queryset = queryset.filter(mode=mode)
    
    # Limit to 10000 records
    queryset = queryset.order_by('-created_at')[:10000]
    
    # Prepare data
    data = []
    for conv in queryset:
        data.append({
            'id': str(conv.id),
            'account_name': conv.account.name if conv.account else '',
            'phone_number': conv.phone_number,
            'contact_name': conv.contact_name,
            'mode': conv.mode,
            'status': conv.status,
            'tags': ', '.join(conv.tags) if conv.tags else '',
            'last_message_at': conv.last_message_at.isoformat() if conv.last_message_at else '',
            'closed_at': conv.closed_at.isoformat() if conv.closed_at else '',
            'resolved_at': conv.resolved_at.isoformat() if conv.resolved_at else '',
            'created_at': conv.created_at.isoformat(),
        })
    
    # Generate response
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    if export_format == 'xlsx':
        return generate_excel_response(data, f'conversations_{timestamp}.xlsx', 'Conversations')
    return generate_csv_response(data, f'conversations_{timestamp}.csv')


